"""
Write rendered document text to .txt, .md, .pdf, or .xlsx format.
"""

import random
from pathlib import Path
from typing import Literal

DocFormat = Literal["txt", "md", "pdf", "xlsx"]

FORMATS: list[DocFormat] = ["txt", "md", "pdf", "xlsx"]
WEIGHTS = [0.30, 0.30, 0.20, 0.20]   # distribution across formats


def pick_format() -> DocFormat:
    return random.choices(FORMATS, weights=WEIGHTS, k=1)[0]


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------

def write_txt(path: Path, text: str) -> None:
    path.write_text(text, encoding="utf-8")


def write_md(path: Path, text: str, facts: dict) -> None:
    company = facts.get("company", "Company")
    category = facts.get("category", "document").replace("_", " ").title()
    policy_key = next((k for k in facts if "policy_id" in k or k == "checklist_id"
                       or k == "course_id" or k == "plan_id"), None)
    policy_id = facts.get(policy_key, "") if policy_key else ""
    header = f"# {company} — {category}\n"
    if policy_id:
        header += f"**Policy ID:** {policy_id}\n\n"
    # Convert plain text paragraphs to markdown paragraphs
    lines = text.split("\n")
    md_lines = []
    for line in lines:
        stripped = line.strip()
        if not stripped:
            md_lines.append("")
        elif stripped.isupper() and len(stripped) > 3:
            md_lines.append(f"## {stripped.title()}")
        else:
            md_lines.append(line)
    path.write_text(header + "\n".join(md_lines), encoding="utf-8")


def _ascii_safe(text: str) -> str:
    """Replace non-latin-1 characters for fpdf2 built-in fonts."""
    return (text
            .replace("—", "--")   # em dash
            .replace("–", "-")    # en dash
            .replace("‘", "'").replace("’", "'")
            .replace("“", '"').replace("”", '"'))


def write_pdf(path: Path, text: str, facts: dict) -> None:
    try:
        from fpdf import FPDF
    except ImportError:
        raise ImportError("fpdf2 required: pip install fpdf2")

    text = _ascii_safe(text)
    company = _ascii_safe(facts.get("company", "Company"))
    category = facts.get("category", "document").replace("_", " ").title()

    pdf = FPDF(format="A4")
    pdf.set_margins(left=20, top=20, right=20)
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()
    body_w = pdf.w - pdf.l_margin - pdf.r_margin

    def _cell(txt: str, bold: bool = False, size: int = 10, h: float = 6) -> None:
        pdf.set_font("Helvetica", "B" if bold else "", size)
        # Encode to latin-1 with replacement to avoid any stray chars
        safe = txt.encode("latin-1", "replace").decode("latin-1")
        pdf.multi_cell(body_w, h, safe)

    _cell(company, bold=True, size=13, h=8)
    _cell(category, bold=True, size=11, h=7)
    pdf.ln(3)

    for line in text.split("\n"):
        stripped = line.strip()
        if not stripped:
            pdf.ln(3)
            continue
        if stripped.isupper() and len(stripped) > 3:
            _cell(stripped, bold=True, h=7)
        else:
            _cell(line.rstrip())

    pdf.output(str(path))


def write_xlsx(path: Path, text: str, facts: dict) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = facts.get("category", "Policy").replace("_", " ").title()

    # Header row
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50

    company = facts.get("company", "Company")
    ws["A1"] = company
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:B1")

    ws["A2"] = facts.get("category", "").replace("_", " ").title()
    ws["A2"].font = Font(bold=True, size=11)
    ws.merge_cells("A2:B2")

    # Key facts table
    skip_keys = {"category", "company"}
    row = 4
    ws[f"A{row}"] = "Field"
    ws[f"B{row}"] = "Value"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].font = Font(bold=True)
    row += 1

    for key, val in facts.items():
        if key in skip_keys:
            continue
        ws[f"A{row}"] = key.replace("_", " ").title()
        ws[f"B{row}"] = str(val)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws[f"B{row}"].alignment = Alignment(wrap_text=True)
        row += 1

    # Full policy text in a separate sheet
    ws2 = wb.create_sheet(title="Full Policy")
    ws2.column_dimensions["A"].width = 100
    for i, line in enumerate(text.split("\n"), start=1):
        ws2[f"A{i}"] = line
        ws2[f"A{i}"].alignment = Alignment(wrap_text=True)

    wb.save(str(path))


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------

def write_doc(output_dir: Path, doc_id: str, text: str, facts: dict,
              fmt: DocFormat | None = None) -> Path:
    """Write document to output_dir/category/doc_id.{fmt}. Returns path written."""
    if fmt is None:
        fmt = pick_format()

    category_dir = output_dir / facts["category"]
    category_dir.mkdir(parents=True, exist_ok=True)

    path = category_dir / f"{doc_id}.{fmt}"

    if fmt == "txt":
        write_txt(path, text)
    elif fmt == "md":
        write_md(path, text, facts)
    elif fmt == "pdf":
        write_pdf(path, text, facts)
    elif fmt == "xlsx":
        write_xlsx(path, text, facts)
    else:
        raise ValueError(f"Unknown format: {fmt}")

    return path
